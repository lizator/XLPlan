from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox

opsumSheets = []
templateData = ["B102", "C102", "D102", "E102", "F102", "G102", "H102", "I102"]
opsummeringData = ["B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5"]
samletData = ["C", "D", "E", "F", "G", "H", "I", "J"]
samletCount = 6

nameOpsum = "Intet dokument valgt"
nameData = "Intet dokument valgt"

sheetCount = 0

directory = "/"

def updateDirectory(filenameList):
    global directory
    directory = ""
    filenameList[len(filenameList) - 1] = None
    for x in filenameList:
        if str(x) != "None":
            directory += str(x) + '/'

def opsumSelect():
    global opsumSheets; global wbOpsum; global nameOpsum; global sheetCount; global samletCount; global directory
    #Choosing Opsum file
    root.filename =  filedialog.askopenfilename(initialdir = directory,title = "Select file",filetypes = (("Excel files","*.xlsx"),("All files","*.*")))
    wbOpsum = load_workbook(root.filename)
    nameOpsumList = root.filename.split("/")
    nameOpsum = nameOpsumList[len(nameOpsumList) - 1]

    updateDirectory(nameOpsumList)

    #Update sheetCount (checking count of allready sheets processed) and samletCount
    for sheet in wbOpsum:
        if (sheet.title != "Samlet" and sheet.title != "Gennemsnit" and sheet.title.find("Template") == -1):
            sheetCount += 1
            samletCount += 1

    #Updates Label
    textOpsum = "Dokument valgt: " + nameOpsum
    vOpsum.set(textOpsum)
    messagebox.showinfo("Reminder!", "HUSK HUSK HUSK HUSK HUSK HUSK!\nHusk at det valgte dokument skal være lukket!")

def dataSelect():
    global opsumSheets; global wbOpsum; global nameOpsum; global nameData; global wsOpsumSamlet; global wsGennemsnit; global wsTemplate; global sheetCount; global opsummeringData; global samletCount; global samletData; global directory
    if (nameOpsum != "Intet dokument valgt"):
        #loading chosen Workbook
        root.filename =  filedialog.askopenfilename(initialdir = directory,title = "Select file",filetypes = (("Excel files","*.xlsx"),("All files","*.*")))
        nameDataList = root.filename.split("/")
        nameData = nameDataList[len(nameDataList) - 1]
        wbData = load_workbook(root.filename, data_only=True)
        wsData = wbData.active

        #Choosing first free sheet
        wsOutput = wbOpsum["Template (" + str(sheetCount + 1) + ")"]

        #choosing 'samlet sheet'
        wsSamlet = wbOpsum['Samlet']
        startCelle = 'C' + str(samletCount); slutCelle = 'J' + str(samletCount)
        samletDataCells = wsSamlet[startCelle:slutCelle] #Belbin numbers

        #Inputting data to wsOutput
        outputDataCells = wsData ["B102":"I102"] #Belbin numbers
        for row in outputDataCells:
            k = 0
            for cell in row:
                wsOutput[opsummeringData[k]].value = cell.value
                wsSamlet[(str(samletData[k]) + str(samletCount))].value = cell.value
                k += 1
        wsOutput.title = wsData["B3"].value #updating opsum sheet title
        wsOutput["B2"].value = wsData["B3"].value #Data Name
        nameCell = 'B' + str(samletCount)
        wsSamlet[nameCell].hyperlink = Hyperlink(nameCell, "'" + wsOutput.title + "'!A1", None, wsData["B3"].value)
        wsSamlet[nameCell].value = wsData["B3"].value

        samletCount += 1 #updating samletCount
        sheetCount += 1 #Updating sheetCount
        wbOpsum.save(directory + nameOpsum)

        #Updates Label
        textData = "Dokument sidst valgt: " + nameData
        vData.set(textData)
    else:
        messagebox.showinfo("Fejl", "Du skal vælge et samle-dokument først :P")

def reset(): #resets opgave er at fjerne alt indsat data fra de enkelte sheets
    global wbOpsum; global nameOpsum; wsSamlet = wbOpsum['Samlet']; global nameData; global sheetCount; global samletCount
    tempCount = 1
    for sheet in wbOpsum:
        if (sheet.title != "Samlet" and sheet.title != "Gennemsnit" and sheet.title.find("Template") == -1):
            for row in sheet["B5":"I5"]:
                for cell in row:
                    cell.value = "None"
            sheet["B3"] = "Name"
            sheet.title = "Template (" + str(tempCount) + ")" #fixing sheet title
            tempCount += 1
    #fixing "Samlet" sheet
    for row in wsSamlet["B6":"J75"]:
        for cell in row:
            cell.value = None

    sheetCount = 0
    samletCount = 6
    nameData = "Intet dokument valgt"
    textData = "Dokument valgt: " + nameData
    vData.set(textData)
    wbOpsum.save(directory + nameOpsum)
    messagebox.showinfo("Succes", "Dokumentet er nu resettet")


root = Tk()
root.title("XL PLAN sortering")

mellemFrame = Frame(root, height=30, width=400)
mellemFrame.pack_propagate(0) # don't shrink
mellemFrame.pack()

opsumFrame = Frame(root, height=80, width=400)
opsumFrame.pack_propagate(0) # don't shrink
opsumFrame.pack()

vOpsum = StringVar()
textOpsum = "Dokument valgt: " + nameOpsum
vOpsum.set(textOpsum)
opsuml1 = Label(opsumFrame, text="Step 1: Vælg samlings-dokumentet")
opsuml2 = Label(opsumFrame, textvariable=vOpsum)
opsumb = Button(opsumFrame, text="Vælg samle-dokument", command=opsumSelect, height=1, width=20)

opsuml1.pack(side=TOP)
opsumb.pack(side=LEFT)
opsuml2.pack(side=RIGHT)

mellemFrame2 = Frame(root, height=30, width=400)
mellemFrame2.pack_propagate(0) # don't shrink
mellemFrame2.pack()

dataFrame = Frame(root, height=70, width=400)
dataFrame.pack_propagate(0) # don't shrink
dataFrame.pack()

vData = StringVar()
textData = "Dokument valgt: " + nameData
vData.set(textData)
datal1 = Label(dataFrame, text="Step 2: Vælg kursisternes dokumenter en efter en")
datal2 = Label(dataFrame, textvariable=vData)
datab = Button(dataFrame, text="Vælg dokument", command=dataSelect, height=1, width=20)

datal1.pack(side=TOP)
datab.pack(side=LEFT)
datal2.pack(side=RIGHT)

mellemFrame3 = Frame(root, height=15, width=400)
mellemFrame3.pack_propagate(0) # don't shrink
mellemFrame3.pack()

resetFrame = Frame(root, height=100, width=400)
resetFrame.pack_propagate(0) # don't shrink
resetFrame.pack()

resetl1 = Label(resetFrame, wraplength=400 ,text="""Programmet vil altid indsætte ny data \
efter allerede indsat data. Hvis noget \
skal laves om, så skal programmet resettes ved \
at trykke på knappen herunder.""")
resetl2 = Label(resetFrame, text="OBS: Dette vil slette alt tidligere data indsat i dokumentet")
resetb = Button(resetFrame, text="RESET", command=reset, height=1, width=20)

resetl1.pack()
resetb.pack(side=BOTTOM)
resetl2.pack(side=BOTTOM)

root.minsize(440,360)

root.mainloop()
